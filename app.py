from flask import Flask, render_template, request, redirect, session, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_wtf.csrf import CSRFProtect, CSRFError
from werkzeug.security import generate_password_hash, check_password_hash
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import urllib.request
import urllib.error
import json
import os

# ── Cloudinary (only active when env vars present) ─────────────────────────────
try:
    import cloudinary
    import cloudinary.uploader
    CLOUDINARY_AVAILABLE = bool(
        os.environ.get('CLOUDINARY_CLOUD_NAME') and
        os.environ.get('CLOUDINARY_API_KEY') and
        os.environ.get('CLOUDINARY_API_SECRET')
    )
    if CLOUDINARY_AVAILABLE:
        cloudinary.config(
            cloud_name = os.environ['CLOUDINARY_CLOUD_NAME'],
            api_key    = os.environ['CLOUDINARY_API_KEY'],
            api_secret = os.environ['CLOUDINARY_API_SECRET'],
        )
except ImportError:
    CLOUDINARY_AVAILABLE = False

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
LOCAL_UPLOAD = os.path.join(BASE_DIR, 'static', 'uploads')

app = Flask(__name__)

# ─── DATABASE ──────────────────────────────────────────────────────────────────
# Render sets DATABASE_URL automatically when a PostgreSQL db is attached.
# Locally falls back to SQLite — no setup needed.
_db_url = os.environ.get('DATABASE_URL', '')
if _db_url.startswith('postgres://'):
    _db_url = _db_url.replace('postgres://', 'postgresql://', 1)
if not _db_url:
    _db_url = 'sqlite:///' + os.path.join(BASE_DIR, 'database.db')

app.config['SQLALCHEMY_DATABASE_URI']        = _db_url
app.config['SQLALCHEMY_ENGINE_OPTIONS']      = {'pool_pre_ping': True}
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# ─── CORE ─────────────────────────────────────────────────────────────────────
app.secret_key = os.environ.get('SECRET_KEY', 'CHANGE-THIS-TO-A-LONG-RANDOM-STRING')
app.config['ALLOWED_EXTENSIONS']         = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
app.config['MAX_CONTENT_LENGTH']         = 5 * 1024 * 1024
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)

db   = SQLAlchemy(app)
csrf = CSRFProtect(app)

ORDER_STAGES = ['Pending', 'Verifying', 'Processing', 'Shipped', 'Completed']


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def allowed_file(filename):
    return (
        isinstance(filename, str) and filename.strip() != '' and
        '.' in filename and
        filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']
    )


def save_image(file_obj, folder='designs'):
    """Upload to Cloudinary in production, local disk in dev.
    Always returns a full URL string, or None on failure."""
    if not file_obj or not allowed_file(file_obj.filename):
        return None

    if CLOUDINARY_AVAILABLE:
        try:
            result = cloudinary.uploader.upload(
                file_obj,
                folder=f"threadline/{folder}",
                resource_type='image',
            )
            return result['secure_url']
        except Exception as e:
            print(f"[CLOUDINARY ERROR] {e}")
            return None
    else:
        from werkzeug.utils import secure_filename
        os.makedirs(LOCAL_UPLOAD, exist_ok=True)
        safe = secure_filename(file_obj.filename)
        if not safe:
            return None
        file_obj.save(os.path.join(LOCAL_UPLOAD, safe))
        return f'/static/uploads/{safe}'


def safe_int(value, default=1, minimum=None, maximum=None):
    try:
        result = int(value)
    except (ValueError, TypeError):
        result = default
    if minimum is not None:
        result = max(minimum, result)
    if maximum is not None:
        result = min(maximum, result)
    return result


def send_email(to, subject, body):
    """Send email via Resend HTTP API — credentials read from Settings DB.
    Configure API key and From address in the admin dashboard Email section."""
    try:
        s = Settings.query.first()
        api_key   = (s.resend_api_key or '').strip() if s else ''
        from_addr = (s.resend_from    or '').strip() if s else ''

        if not api_key or not from_addr:
            print("[EMAIL] Resend not configured in dashboard — skipping.")
            return

        payload = json.dumps({
            "from":    from_addr,
            "to":      [to],
            "subject": subject,
            "text":    body,
        }).encode('utf-8')

        req = urllib.request.Request(
            "https://api.resend.com/emails",
            data    = payload,
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type":  "application/json",
            },
            method = "POST",
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            print(f"[EMAIL] Sent to {to} — status {resp.status}")
    except urllib.error.HTTPError as e:
        print(f"[EMAIL ERROR] HTTP {e.code}: {e.read().decode()}")
    except Exception as e:
        print(f"[EMAIL ERROR] {e}")


def send_admin_alert(subject, body):
    try:
        s = Settings.query.first()
        if s and s.admin_email and s.admin_email.strip():
            send_email(s.admin_email.strip(), subject, body)
    except Exception as e:
        print(f"[ADMIN ALERT ERROR] {e}")


def build_invoice_pdf(order_id):
    """Generate invoice entirely in memory — no disk write needed."""
    o = db.session.get(Order, order_id)
    if not o:
        return None
    d     = Design.query.filter_by(design_code=o.design_code).first()
    price = d.price if d else '—'

    buf = BytesIO()
    c   = canvas.Canvas(buf, pagesize=letter)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(180, 760, "THREADLINE — INVOICE")
    c.setFont("Helvetica", 12)
    for y, text in [
        (720, f"Order ID  : #{o.id}"),
        (700, f"Date      : {datetime.now().strftime('%d-%m-%Y')}"),
        (670, f"Customer  : {o.customer_name}"),
        (650, f"Phone     : {o.phone}"),
        (630, f"Address   : {o.house}, {o.city}, {o.mandal} - {o.pincode}"),
        (600, f"Product   : {o.design}"),
        (580, f"Size      : {o.size}"),
        (560, f"Quantity  : {o.quantity}"),
        (540, f"Amount    : Rs.{price}"),
        (518,  "Estimated Delivery : 3-5 business days"),
    ]:
        c.drawString(50, y, text)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, 490, "Thank you for shopping with THREADLINE!")
    c.save()
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
#  MODELS
# ══════════════════════════════════════════════════════════════════════════════

class Settings(db.Model):
    id               = db.Column(db.Integer, primary_key=True)
    admin_password   = db.Column(db.String(200), default="")
    phonepe_name     = db.Column(db.String(100), default="Your Name")
    phonepe_number   = db.Column(db.String(20),  default="9999999999")
    admin_whatsapp   = db.Column(db.String(20),  default="")
    # Resend email config — set from admin dashboard
    resend_api_key   = db.Column(db.String(200), default="")   # re_xxxxxxxxxxxx
    resend_from      = db.Column(db.String(200), default="")   # THREADLINE <orders@yourdomain.com>
    admin_email      = db.Column(db.String(200), default="")   # where YOU receive order alerts
    # legacy columns kept so existing DBs don't break (unused)
    mail_username    = db.Column(db.String(200), default="")
    mail_password    = db.Column(db.String(200), default="")


class Design(db.Model):
    id             = db.Column(db.Integer, primary_key=True)
    design_code    = db.Column(db.String(20), unique=True)
    name           = db.Column(db.String(100))
    description    = db.Column(db.String(300))
    price          = db.Column(db.String(20))
    image          = db.Column(db.String(500))   # full URL
    stock          = db.Column(db.String(20), default="In Stock")
    stock_quantity = db.Column(db.Integer,    default=10)
    images         = db.relationship(
        'DesignImage', backref='design',
        cascade='all, delete-orphan',
        order_by='DesignImage.sort_order'
    )


class DesignImage(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    design_id  = db.Column(db.Integer, db.ForeignKey('design.id'), nullable=False)
    filename   = db.Column(db.String(500), nullable=False)   # full URL
    sort_order = db.Column(db.Integer, default=0)


class Order(db.Model):
    id                = db.Column(db.Integer, primary_key=True)
    design            = db.Column(db.String(100))
    design_code       = db.Column(db.String(20))
    customer_name     = db.Column(db.String(100))
    house             = db.Column(db.String(100))
    city              = db.Column(db.String(100))
    mandal            = db.Column(db.String(100))
    pincode           = db.Column(db.String(10),  default="")
    email             = db.Column(db.String(100))
    size              = db.Column(db.String(10))
    quantity          = db.Column(db.String(10))
    phone             = db.Column(db.String(20))
    payment_image     = db.Column(db.String(500))   # full URL
    status            = db.Column(db.String(20),  default="Pending")
    created_at        = db.Column(db.String(50),  nullable=True)
    completed_at      = db.Column(db.String(50),  nullable=True)
    cancelled_at      = db.Column(db.String(50),  nullable=True)
    status_updated_at = db.Column(db.String(50),  nullable=True)


with app.app_context():
    db.create_all()
    if not Settings.query.first():
        db.session.add(Settings(admin_password=generate_password_hash("admin123")))
        db.session.commit()
    if not CLOUDINARY_AVAILABLE:
        os.makedirs(LOCAL_UPLOAD, exist_ok=True)
    # Idempotent seed: DesignImage from Design.image for legacy rows
    for dsg in Design.query.all():
        if dsg.image and not DesignImage.query.filter_by(design_id=dsg.id, sort_order=0).first():
            db.session.add(DesignImage(design_id=dsg.id, filename=dsg.image, sort_order=0))
    db.session.commit()


# ══════════════════════════════════════════════════════════════════════════════
#  ERROR HANDLERS
# ══════════════════════════════════════════════════════════════════════════════

@app.errorhandler(404)
def page_not_found(e):
    return render_template("404.html"), 404

@app.errorhandler(413)
def file_too_large(e):
    return render_template("404.html", size_error=True), 413

@app.errorhandler(500)
def internal_error(e):
    db.session.rollback()
    return render_template("404.html", server_error=True), 500

@app.errorhandler(CSRFError)
def csrf_error(e):
    # CSRF token expired or missing — send user back to login
    session.clear()
    return redirect('/admin')


# ══════════════════════════════════════════════════════════════════════════════
#  PUBLIC ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/')
def home():
    return render_template("home.html",
        designs=Design.query.all(),
        settings=Settings.query.first())


@app.route('/track', methods=['GET', 'POST'])
def track_order():
    orders = []; searched = False; phone = ''
    if request.method == 'POST':
        phone = request.form.get('phone', '').strip()
        searched = True
        if phone:
            orders = Order.query.filter_by(phone=phone).order_by(Order.id.desc()).all()
    return render_template("track_order.html", orders=orders, searched=searched, phone=phone)


@app.route('/order/<design_code>', methods=['GET', 'POST'])
def order(design_code):
    settings = Settings.query.first()
    design   = Design.query.filter_by(design_code=design_code).first()
    if not design or design.stock == "Out of Stock":
        return redirect('/')

    if request.method == 'POST':
        phone = request.form.get('phone', '').strip()

        if Order.query.filter_by(phone=phone, design_code=design_code, status='Pending').first():
            return render_template("order.html", design=design, settings=settings,
                                   duplicate=True, file_error=None)

        payment = request.files.get('payment')
        if not payment or not allowed_file(payment.filename):
            return render_template("order.html", design=design, settings=settings,
                                   duplicate=False,
                                   file_error="Only PNG / JPG / JPEG / GIF / WEBP images are allowed.")

        payment_url = save_image(payment, folder='payments')
        if not payment_url:
            return render_template("order.html", design=design, settings=settings,
                                   duplicate=False,
                                   file_error="Upload failed. Please try a different image file.")

        qty = safe_int(request.form.get('quantity', 1), default=1, minimum=1)

        new_order = Order(
            design        = design.name,
            design_code   = design.design_code,
            customer_name = request.form.get('customer_name', '').strip(),
            house         = request.form.get('house', '').strip(),
            city          = request.form.get('city', '').strip(),
            mandal        = request.form.get('mandal', '').strip(),
            pincode       = request.form.get('pincode', '').strip(),
            email         = request.form.get('email', '').strip(),
            size          = request.form.get('size', 'M'),
            quantity      = str(qty),
            phone         = phone,
            payment_image = payment_url,
            created_at    = datetime.now().strftime('%d-%m-%Y %I:%M %p'),
        )
        db.session.add(new_order)
        design.stock_quantity = max(0, design.stock_quantity - qty)
        if design.stock_quantity == 0:
            design.stock = "Out of Stock"
        db.session.commit()

        send_email(new_order.email,
            f"Order Confirmed #{new_order.id} | THREADLINE",
            f"Hi {new_order.customer_name},\n\nYour order has been placed!\n\n"
            f"Order ID : #{new_order.id}\nDesign   : {new_order.design}\n"
            f"Size     : {new_order.size}  |  Qty: {new_order.quantity}\n"
            f"Amount   : Rs.{design.price}\nPlaced   : {new_order.created_at}\n\n"
            f"Address  : {new_order.house}, {new_order.city}, {new_order.mandal} - {new_order.pincode}\n\n"
            f"Estimated Delivery: 3-5 business days after payment verification.\n"
            f"Track your order: {request.host_url}track\n\nThank you — THREADLINE Team")

        send_admin_alert(
            f"New Order #{new_order.id} — {new_order.customer_name}",
            f"New order on THREADLINE!\n\nOrder #{new_order.id}\n"
            f"Customer : {new_order.customer_name}\nPhone    : {new_order.phone}\n"
            f"Email    : {new_order.email}\nDesign   : {new_order.design} ({new_order.size}) x{qty}\n"
            f"Amount   : Rs.{design.price}\n"
            f"Address  : {new_order.house}, {new_order.city}, {new_order.mandal} - {new_order.pincode}\n"
            f"Time     : {new_order.created_at}\n\nGo to dashboard to update the order status.")

        return render_template("success.html", order_id=new_order.id)

    return render_template("order.html", design=design, settings=settings,
                           duplicate=False, file_error=None)


@app.route('/invoice/<int:order_id>')
def download_invoice(order_id):
    """Generate invoice in-memory on demand — no disk storage needed."""
    buf = build_invoice_pdf(order_id)
    if not buf:
        return redirect('/')
    return send_file(buf, download_name=f"invoice_{order_id}.pdf",
                     as_attachment=True, mimetype='application/pdf')


# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN AUTH
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/admin', methods=['GET', 'POST'])
def admin():
    settings = Settings.query.first()
    error = None
    if request.method == 'POST':
        if settings and check_password_hash(settings.admin_password,
                                            request.form.get('password', '')):
            session.permanent = True
            session['admin']  = True
            return redirect('/dashboard')
        error = "Incorrect password. Please try again."
    return render_template("login.html", error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')


@app.route('/change_password', methods=['GET', 'POST'])
def change_password():
    if not session.get('admin'):
        return redirect('/admin')
    s = Settings.query.first()
    if request.method == 'POST':
        new_pw = request.form.get('new_password', '').strip()
        if new_pw:
            s.admin_password = generate_password_hash(new_pw)
            db.session.commit()
        return redirect('/dashboard')
    return render_template("change_password.html")


# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/dashboard')
def dashboard():
    if not session.get('admin'):
        return redirect('/admin')

    active_orders    = Order.query.filter(
        Order.status.notin_(['Completed', 'Cancelled'])
    ).order_by(Order.id.desc()).all()
    completed_orders = Order.query.filter_by(status='Completed').order_by(Order.id.desc()).all()
    cancelled_orders = Order.query.filter_by(status='Cancelled').order_by(Order.id.desc()).all()
    designs          = Design.query.all()
    settings         = Settings.query.first()

    total_revenue = 0; monthly_revenue = 0
    current_month = datetime.now().strftime('%m-%Y')
    for o in completed_orders:
        d = Design.query.filter_by(design_code=o.design_code).first()
        if d:
            try:
                amt = float(d.price) * safe_int(o.quantity, default=1, minimum=0)
                total_revenue += amt
                if o.completed_at and current_month in o.completed_at:
                    monthly_revenue += amt
            except Exception:
                pass

    best_design = "N/A"; best_count = 0
    for d in designs:
        cnt = Order.query.filter_by(design_code=d.design_code).count()
        if cnt > best_count:
            best_count = cnt; best_design = d.name

    return render_template("dashboard.html",
        active_orders=active_orders, completed_orders=completed_orders,
        cancelled_orders=cancelled_orders, designs=designs, settings=settings,
        total_revenue=int(total_revenue), monthly_revenue=int(monthly_revenue),
        best_design=best_design, order_stages=ORDER_STAGES)


@app.route('/update_status/<int:order_id>', methods=['POST'])
def update_status(order_id):
    if not session.get('admin'):
        return redirect('/admin')
    o = db.session.get(Order, order_id)
    if not o:
        return redirect('/dashboard')

    new_status = request.form.get('status', '').strip()
    if new_status not in ORDER_STAGES + ['Cancelled']:
        return redirect('/dashboard')

    old_status = o.status
    now        = datetime.now().strftime('%d-%m-%Y %I:%M %p')

    if new_status == 'Completed':
        o.completed_at = now
    elif new_status == 'Cancelled':
        o.cancelled_at = now
        d = Design.query.filter_by(design_code=o.design_code).first()
        if d:
            try:
                d.stock_quantity += safe_int(o.quantity, default=1, minimum=0)
                d.stock = "In Stock"
            except Exception:
                pass
    else:
        o.status_updated_at = now
        o.completed_at = None
        o.cancelled_at = None

    o.status = new_status
    db.session.commit()

    if new_status != old_status:
        msgs = {
            'Verifying':  ("Payment Verification — THREADLINE",
                f"Hi {o.customer_name},\n\nWe received your payment screenshot and are verifying it.\n\n"
                f"Order #{o.id} — {o.design} ({o.size} x{o.quantity})\n\nThank you — THREADLINE Team"),
            'Processing': ("Order Being Processed — THREADLINE",
                f"Hi {o.customer_name},\n\nPayment verified! Your order is now being printed.\n\n"
                f"Order #{o.id} — {o.design} ({o.size} x{o.quantity})\n\nThank you — THREADLINE Team"),
            'Shipped':    ("Your Order Is On Its Way — THREADLINE",
                f"Hi {o.customer_name},\n\nYour order has been shipped!\n\n"
                f"Order #{o.id} — {o.design} ({o.size} x{o.quantity})\n"
                f"Estimated delivery: 3-5 business days.\n\nThank you — THREADLINE Team"),
            'Completed':  ("Order Delivered — THREADLINE",
                f"Hi {o.customer_name},\n\nYour order has been delivered. We hope you love it!\n\n"
                f"Order #{o.id} — {o.design}\n\nThank you for shopping with THREADLINE!"),
            'Cancelled':  ("Order Cancelled — THREADLINE",
                f"Hi {o.customer_name},\n\nOrder #{o.id} ({o.design}) has been cancelled.\n"
                f"If this was a mistake, please contact us.\n\nThank you — THREADLINE Team"),
        }
        if new_status in msgs:
            send_email(o.email, *msgs[new_status])

    return redirect('/dashboard')


@app.route('/update_phonepe', methods=['POST'])
def update_phonepe():
    if not session.get('admin'):
        return redirect('/admin')
    s = Settings.query.first()
    s.phonepe_name   = request.form.get('name', '').strip()
    s.phonepe_number = request.form.get('number', '').strip()
    s.admin_whatsapp = request.form.get('whatsapp', '').strip().replace('+', '').replace(' ', '')
    db.session.commit()
    return redirect('/dashboard')


@app.route('/update_email_config', methods=['POST'])
def update_email_config():
    if not session.get('admin'):
        return redirect('/admin')
    s = Settings.query.first()
    # Only update API key if a new one was typed (blank = keep existing)
    new_key = request.form.get('resend_api_key', '').strip()
    if new_key:
        s.resend_api_key = new_key
    s.resend_from  = request.form.get('resend_from',  '').strip()
    s.admin_email  = request.form.get('admin_email',  '').strip()
    db.session.commit()
    return redirect('/dashboard')


@app.route('/send_test_email', methods=['POST'])
def send_test_email():
    """Send a test email to the admin_email address to verify config works."""
    if not session.get('admin'):
        return redirect('/admin')
    s = Settings.query.first()
    test_to = (s.admin_email or '').strip() if s else ''
    if not test_to:
        return redirect('/dashboard')
    send_email(
        test_to,
        "✅ THREADLINE — Test Email",
        "This is a test email from your THREADLINE admin dashboard.\n\n"
        "If you received this, your Resend email configuration is working correctly!\n\n"
        "— THREADLINE Admin"
    )
    return redirect('/dashboard')


# ── DESIGN MANAGEMENT ─────────────────────────────────────────────────────────

@app.route('/add_design', methods=['GET', 'POST'])
def add_design():
    if not session.get('admin'):
        return redirect('/admin')
    error = None
    if request.method == 'POST':
        images       = request.files.getlist('images')
        valid_images = [f for f in images if f and allowed_file(f.filename)]
        if not valid_images:
            error = "Please upload at least one image (PNG, JPG, JPEG, GIF, WEBP)."
        else:
            try:
                qty       = safe_int(request.form.get('stock_quantity', 10), default=10, minimum=0)
                first_url = save_image(valid_images[0], folder='designs')
                if not first_url:
                    error = "Failed to upload the image. Please try again."
                else:
                    new_design = Design(
                        design_code    = request.form.get('code', '').strip(),
                        name           = request.form.get('name', '').strip(),
                        description    = request.form.get('description', '').strip(),
                        price          = request.form.get('price', '0').strip(),
                        image          = first_url,
                        stock          = "In Stock" if qty > 0 else "Out of Stock",
                        stock_quantity = qty,
                    )
                    db.session.add(new_design)
                    db.session.flush()
                    for i, img_file in enumerate(valid_images):
                        url = first_url if i == 0 else save_image(img_file, folder='designs')
                        if url:
                            db.session.add(DesignImage(
                                design_id=new_design.id, filename=url, sort_order=i))
                    db.session.commit()
                    return redirect('/dashboard')
            except Exception as e:
                db.session.rollback()
                print(f"[ADD DESIGN ERROR] {e}")
                error = "Something went wrong while saving. Please try again."
    return render_template("add_design.html", error=error)


@app.route('/edit_design/<int:design_id>', methods=['GET', 'POST'])
def edit_design(design_id):
    if not session.get('admin'):
        return redirect('/admin')
    d = db.session.get(Design, design_id)
    if not d:
        return redirect('/dashboard')
    error = None
    if request.method == 'POST':
        try:
            d.name           = request.form.get('name', d.name).strip()
            d.description    = request.form.get('description', d.description).strip()
            d.price          = request.form.get('price', d.price).strip()
            d.stock_quantity = safe_int(request.form.get('stock_quantity', d.stock_quantity),
                                        default=d.stock_quantity, minimum=0)
            d.stock = "In Stock" if d.stock_quantity > 0 else "Out of Stock"

            new_images = request.files.getlist('images')
            valid_new  = [f for f in new_images if f and allowed_file(f.filename)]
            if valid_new:
                max_order = db.session.query(db.func.max(DesignImage.sort_order))\
                              .filter_by(design_id=d.id).scalar() or -1
                for i, img_file in enumerate(valid_new):
                    url = save_image(img_file, folder='designs')
                    if url:
                        db.session.add(DesignImage(
                            design_id=d.id, filename=url, sort_order=max_order + 1 + i))

            first = DesignImage.query.filter_by(design_id=d.id)\
                      .order_by(DesignImage.sort_order).first()
            if first:
                d.image = first.filename
            db.session.commit()
            return redirect('/dashboard')
        except Exception as e:
            db.session.rollback()
            print(f"[EDIT DESIGN ERROR] {e}")
            error = "Something went wrong while saving. Please try again."
    return render_template("edit_design.html", design=d, error=error)


@app.route('/delete_design_image/<int:image_id>')
def delete_design_image(image_id):
    if not session.get('admin'):
        return redirect('/admin')
    img = db.session.get(DesignImage, image_id)
    if not img:
        return redirect('/dashboard')
    design_id = img.design_id
    if DesignImage.query.filter_by(design_id=design_id).count() > 1:
        db.session.delete(img)
        new_first = DesignImage.query.filter_by(design_id=design_id)\
                      .order_by(DesignImage.sort_order).first()
        if new_first:
            dsg = db.session.get(Design, design_id)
            if dsg:
                dsg.image = new_first.filename
        db.session.commit()
    return redirect(f'/edit_design/{design_id}')


@app.route('/delete_design/<int:design_id>')
def delete_design(design_id):
    if not session.get('admin'):
        return redirect('/admin')
    d = db.session.get(Design, design_id)
    if d:
        db.session.delete(d); db.session.commit()
    return redirect('/dashboard')


@app.route('/toggle_stock/<int:design_id>')
def toggle_stock(design_id):
    if not session.get('admin'):
        return redirect('/admin')
    d = db.session.get(Design, design_id)
    if d:
        d.stock = "Out of Stock" if d.stock == "In Stock" else "In Stock"
        db.session.commit()
    return redirect('/dashboard')


@app.route('/export_orders')
def export_orders():
    if not session.get('admin'):
        return redirect('/admin')
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Orders"
    headers = ['ID','Design','Code','Customer','House','City','Mandal',
               'Pincode','Phone','Email','Size','Qty','Status',
               'Created At','Completed At','Cancelled At']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF6B00")
        cell.alignment = Alignment(horizontal="center")
    for o in Order.query.order_by(Order.id.desc()).all():
        ws.append([o.id, o.design, o.design_code, o.customer_name,
                   o.house or '', o.city or '', o.mandal or '',
                   o.pincode or '', o.phone or '', o.email or '',
                   o.size or '', o.quantity or '', o.status or '',
                   o.created_at or '', o.completed_at or '', o.cancelled_at or ''])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or '')) for c in col) + 4, 40)
    output = BytesIO(); wb.save(output); output.seek(0)
    return send_file(output,
        download_name=f"orders_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx",
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/sales_analysis')
def sales_analysis():
    if not session.get('admin'):
        return redirect('/admin')
    designs = Design.query.all()
    return render_template("sales_report.html",
        labels=[d.name for d in designs],
        values=[Order.query.filter_by(design_code=d.design_code).count() for d in designs])


if __name__ == "__main__":
    app.run(debug=os.environ.get('FLASK_DEBUG', '0') == '1')
